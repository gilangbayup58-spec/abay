import openpyxl
from datetime import datetime
import os

class AttendanceManager:

    def add_student(self, name):
        raise NotImplementedError
   
def view_attendance(self, name=None):
    if self.ws is None:
        print("Error: Worksheet not loaded. Check the file path and initialization.")
        return
    if name:
        if name not in self.students:
            print(f"Student {name} not found.")
            return
        row = self.students.index(name) + 2
        print(f"Attendance for {name}:")
        for col in range(2, self.ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col) # type: ignore
            date = self.ws[f'{col_letter}1'].value
            status = self.ws[f'{col_letter}{row}'].value
            print(f"{date}: {status}")
    else:
        print("All attendance:")
        for row in range(2, self.ws.max_row + 1):
            name = self.ws[f'A{row}'].value
            print(f"\n{name}:")
            for col in range(2, self.ws.max_column + 1):
                col_letter = openpyxl.utils.get_column_letter(col) # pyright: ignore[reportAttributeAccessIssue]
                date = self.ws[f'{col_letter}1'].value
                status = self.ws[f'{col_letter}{row}'].value
                print(f"  {date}: {status}")

    def save(self):
        self.wb.save(self.filename)

def main():
    manager = AttendanceManager()
    while True:
        print("\n1. Add Student")
        print("2. Mark Attendance")
        print("3. View Attendance")
        print("4. Exit")
        choice = input("Choose an option: ")
        if choice == '1':
            name = input("Enter student name: ")
            manager.add_student(name)
            print(f"Added {name}")
        elif choice == '2':
            name = input("Enter student name: ")
            date_str = input("Enter date (YYYY-MM-DD): ")
            try:
                date = datetime.strptime(date_str, '%Y-%m-%d')
                status = input("Enter status (P/A): ").upper()
                if status in ['P', 'A']:
                    manager.mark_attendance(name, date, status) # type: ignore
                    print("Attendance marked")
                else:
                    print("Invalid status")
            except ValueError:
                print("Invalid date format")
        elif choice == '3':
            name = input("Enter student name (or leave blank for all): ")
            manager.view_attendance(name if name else None) # type: ignore
        elif choice == '4':
            break
        else:
            print("Invalid choice")

if __name__ == "__main__":
    main()